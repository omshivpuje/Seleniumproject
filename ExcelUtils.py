import openpyxl


def load_workbook(path="C://Users//hp//Desktop//pytest.xlsx"):
    return openpyxl.load_workbook(path)


def save_workbook(path="C://Users//hp//Desktop//pytest.xlsx"):
    workbook.save(path)


def get_sheet(sheet_name):
    return workbook.active if sheet_name == "active" else workbook.get_sheet_by_name(sheet_name)


def get_rows_from_sheet(sheet_name="active"):
    sheet = get_sheet(sheet_name)
    return sheet.max_row


def get_columns_from_sheet(sheet_name="active"):
    sheet = get_sheet(sheet_name)
    return sheet.max_column


def get_cell_value(r, c, sheet_name="active"):
    sheet = get_sheet(sheet_name)
    return sheet.cell(row=r, column=c).value


def write_cell_value(r, c, val, sheet_name='active'):
    sheet = get_sheet(sheet_name)
    sheet.cell(row=r, column=c).value = val
    save_workbook()


workbook = load_workbook()
